#!/usr/bin/env python3
"""Genereerib Eesti Kanepimajanduse 5-aasta finants- ja strateegiamudeli Exceli."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/workspace/Eesti-Kanepimajandus-5a-Finantsmudel.xlsx"

# Värvid
GREEN = PatternFill("solid", fgColor="2E7D32")
LIGHT_GREEN = PatternFill("solid", fgColor="E8F5E9")
GOLD = PatternFill("solid", fgColor="FFF8E1")
BLUE = PatternFill("solid", fgColor="E3F2FD")
GRAY = PatternFill("solid", fgColor="F5F5F5")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1B5E20")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, cols, fill=GREEN):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_table(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0:
                cell.fill = GRAY


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title(ws, title, merge_cols=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = LIGHT_GREEN
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


wb = Workbook()

# ─────────────────────────────────────────────
# LEHT 1: KOKKUVÕTE
# ─────────────────────────────────────────────
ws = wb.active
ws.title = "Kokkuvõte"
add_title(ws, "🇪🇪 EESTI KANEPIMAJANDUS 2030 — RIIGI ETTEVÕTTE MUDEL", 8)
set_widths(ws, [28, 14, 14, 14, 14, 14, 14, 36])

rows = [
    ("", "Aasta 1", "Aasta 2", "Aasta 3", "Aasta 4", "Aasta 5", "5a KOKKU", "Märkus"),
    ("TULUD kokku (€)", 1_200_000, 4_500_000, 12_000_000, 22_000_000, 35_000_000, "=SUM(B3:F3)", "Riigiettevõte + partnerid"),
    ("KULUD kokku (€)", 2_800_000, 3_600_000, 6_500_000, 9_000_000, 12_000_000, "=SUM(B4:F4)", "Sh toetused, liising, palgad"),
    ("NETOKASUM (€)", "=B3-B4", "=C3-C4", "=D3-D4", "=E3-E4", "=F3-F4", "=SUM(B5:F5)", "Enne fondi jaotust"),
    ("Riigikassasse (€)", 0, 350_000, 1_800_000, 4_200_000, 7_500_000, "=SUM(B6:F6)", "Maksud + dividendid"),
    ("Uued töökohad", 45, 120, 280, 450, 650, "=SUM(B7:F7)", "Otsetöökohti"),
    ("Puukoolid partnerid", 3, 8, 15, 22, 30, "=SUM(B8:F8)", "Toetust saavad"),
    ("Liisingumasinad (tk)", 4, 12, 28, 45, 60, "=SUM(B9:F9)", "Kliendi maksab kinni"),
    ("", "", "", "", "", "", "", ""),
    ("ETTEVÕTE", "Eesti Roheline Majandus AS (ERMA)", "", "", "", "", "", "Riigi 51% + kasvatajate kooperatiiv 49%"),
    ("MUDEL", "Vaba kassa pärast liisingu lõppu", "", "", "", "", "", "Masinad → riigi/vara, kasum reinvesteeritakse"),
    ("BRÜSSELL", "Tööstuskanep + siseriiklik reguleerimine", "", "", "", "", "", "Saksamaa mudel, CAP, Roheline kokkulepe"),
]

for i, row in enumerate(rows, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)
style_header(ws, 3, 8)
style_table(ws, 4, 11, 8)
for col in "BCDEFG":
    for r in range(4, 10):
        ws[f"{col}{r}"].number_format = '#,##0'

# ─────────────────────────────────────────────
# LEHT 2: TULUD
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Tulud")
add_title(ws2, "TULUD — tuluvoogude detail (€/aasta)", 7)
set_widths(ws2, [32, 14, 14, 14, 14, 14, 40])

tulud = [
    ("Tuluvoog", "A1", "A2", "A3", "A4", "A5", "Kirjeldus"),
    ("Tööstuskanep — kiud ja seemned", 400_000, 1_200_000, 3_000_000, 4_500_000, 6_000_000, "Eksport + kohalik tööstus"),
    ("CBD kosmeetika (EL lubatud)", 150_000, 600_000, 1_500_000, 2_500_000, 4_000_000, "Kreemid, õlid, seebid"),
    ("Riiklik growshop võrgustik", 0, 800_000, 2_500_000, 5_000_000, 8_000_000, "Asendab välismaa poode"),
    ("Reguleeritud täiskasvanute turg", 0, 0, 1_000_000, 5_000_000, 10_000_000, "Aasta 3 piloot, A4-A5 täismahus"),
    ("Meditsiiniline kanep", 100_000, 400_000, 1_200_000, 2_000_000, 3_000_000, "Kohalik tootmine, haigekassa"),
    ("Puukoolide seemikumüük", 80_000, 250_000, 600_000, 1_000_000, 1_500_000, "Certified seedlings"),
    ("Koolitus ja konsultatsioon", 50_000, 150_000, 300_000, 500_000, 800_000, "EPM, kutsekoolid"),
    ("Turism ja kogemuskeskused", 20_000, 100_000, 400_000, 1_500_000, 2_700_000, "Farm visits, loodus"),
    ("Liisingutasud (masinate tagasimaks)", 400_000, 1_000_000, 2_000_000, 3_000_000, 3_500_000, "Klient maksab masina kinni"),
    ("KOKKU", "=SUM(B2:B10)", "=SUM(C2:C10)", "=SUM(D2:D10)", "=SUM(E2:E10)", "=SUM(F2:F10)", ""),
]
for i, row in enumerate(tulud, start=3):
    for j, val in enumerate(row, start=1):
        ws2.cell(row=i, column=j, value=val)
style_header(ws2, 3, 7)
style_table(ws2, 4, 12, 7)
for r in range(4, 13):
    for c in range(2, 7):
        ws2.cell(row=r, column=c).number_format = '#,##0'

# ─────────────────────────────────────────────
# LEHT 3: KULUD
# ─────────────────────────────────────────────
ws3 = wb.create_sheet("Kulud")
add_title(ws3, "KULUD — kulustruktuur (€/aasta)", 7)
set_widths(ws3, [32, 14, 14, 14, 14, 14, 40])

kulud = [
    ("Kulukategooria", "A1", "A2", "A3", "A4", "A5", "Kirjeldus"),
    ("Töötajate palgad", 900_000, 1_400_000, 2_500_000, 3_800_000, 5_200_000, "Keskmine 1800-2200 bruto"),
    ("Liisingumaksed (masinad)", 480_000, 720_000, 1_200_000, 1_500_000, 1_200_000, "Väheneb kui masinad paid off"),
    ("Toetused puukoolidele", 300_000, 450_000, 600_000, 500_000, 400_000, "Seemik, sertifitseerimine"),
    ("Toetused kasvatajatele", 250_000, 400_000, 700_000, 800_000, 900_000, "Hektaripõhine boonus"),
    ("Turustus ja growshopid", 200_000, 350_000, 800_000, 1_200_000, 1_800_000, "Rendid, sisustus, IT"),
    ("Uurimine ja kvaliteet", 150_000, 200_000, 350_000, 400_000, 500_000, "Labor, sertifikaadid"),
    ("Juriidika ja compliance", 120_000, 150_000, 200_000, 250_000, 300_000, "EL reeglid, audit"),
    ("Koolitusprogrammid", 80_000, 150_000, 300_000, 400_000, 500_000, "Koolid, kutsekvalifikatsioon"),
    ("Halduskulud", 200_000, 250_000, 350_000, 450_000, 550_000, "IT, kontor, juhtimine"),
    ("Investeeringud (capex)", 220_000, 430_000, 500_000, 800_000, 650_000, "Päikesepargid, hooned"),
    ("KOKKU", "=SUM(B2:B11)", "=SUM(C2:C11)", "=SUM(D2:D11)", "=SUM(E2:E11)", "=SUM(F2:F11)", ""),
]
for i, row in enumerate(kulud, start=3):
    for j, val in enumerate(row, start=1):
        ws3.cell(row=i, column=j, value=val)
style_header(ws3, 3, 7)
style_table(ws3, 4, 13, 7)
for r in range(4, 14):
    for c in range(2, 7):
        ws3.cell(row=r, column=c).number_format = '#,##0'

# ─────────────────────────────────────────────
# LEHT 4: KASUMIARUANNE
# ─────────────────────────────────────────────
ws4 = wb.create_sheet("Kasumiaruanne")
add_title(ws4, "KASUMIARUANNE (P&L)", 7)
set_widths(ws4, [30, 14, 14, 14, 14, 14, 30])

pl = [
    ("", "A1", "A2", "A3", "A4", "A5", ""),
    ("Tulud kokku", 1_200_000, 4_500_000, 12_000_000, 22_000_000, 35_000_000, ""),
    ("Kulud kokku", 2_800_000, 3_600_000, 6_500_000, 9_000_000, 12_000_000, ""),
    ("EBITDA", "=B4-B5", "=C4-C5", "=D4-D5", "=E4-E5", "=F4-F5", ""),
    ("EBITDA %", "=B6/B4", "=C6/C4", "=D6/D4", "=E6/E4", "=F6/F4", ""),
    ("", "", "", "", "", "", ""),
    ("Jaotus Eesti Tuleviku Fondi:", "", "", "", "", "", ""),
    ("  Päikesepargid (40%)", "=MAX(0,B6*0.4)", "=MAX(0,C6*0.4)", "=D6*0.4", "=E6*0.4", "=F6*0.4", "Macro-kott"),
    ("  Tööstus + robotid (30%)", "=MAX(0,B6*0.3)", "=MAX(0,C6*0.3)", "=D6*0.3", "=E6*0.3", "=F6*0.3", "Macro-kott"),
    ("  Basic income pilot (20%)", "=MAX(0,B6*0.2)", "=MAX(0,C6*0.2)", "=D6*0.2", "=E6*0.2", "=F6*0.2", "Macro-kott"),
    ("  Uurimine + koolid (10%)", "=MAX(0,B6*0.1)", "=MAX(0,C6*0.1)", "=D6*0.1", "=E6*0.1", "=F6*0.1", "Macro-kott"),
    ("", "", "", "", "", "", ""),
    ("Dalio põhimõte", "18 mitte-korrelatsioonset voogu", "", "", "", "", "Vt leht Finantsstrateegia"),
    ("", "", "", "", "", "", ""),
    ("5-aasta kumulatiivne EBITDA", "=SUM(B6:F6)", "", "", "", "", "Eesmärk: positiivne A2-st"),
]
for i, row in enumerate(pl, start=3):
    for j, val in enumerate(row, start=1):
        ws4.cell(row=i, column=j, value=val)
style_header(ws4, 3, 6)
style_table(ws4, 4, 15, 6)
for r in [4, 5, 6, 9, 10, 11, 12, 15]:
    for c in range(2, 7):
        ws4.cell(row=r, column=c).number_format = '#,##0'
for r in [7]:
    for c in range(2, 7):
        ws4.cell(row=r, column=c).number_format = '0.0%'

# ─────────────────────────────────────────────
# LEHT 5: 5-AASTA PLAAN
# ─────────────────────────────────────────────
ws5 = wb.create_sheet("5a arenguplaan")
add_title(ws5, "5-AASTA ARENGUPLAAN 2026–2030", 5)
set_widths(ws5, [12, 28, 28, 28, 40])

plan = [
    ("Aasta", "Fookus", "Tegevused", "Eesmärgid", "Mõõdikud"),
    ("2026 A1", "Vundament", "ERMA AS loomine; 3 puukooli piloot; tööstuskanep klaster Jõgeval; Brüsseli positsioonipaber", "Legaalne tööstus käimas", "45 töökohta, 400k tulu"),
    ("2027 A2", "Kasv", "8 puukooli; 5 growshopi; CBD kosmeetika bränd; kutsekooli programm", "Kohalik toode turul", "120 töökohta, 4.5M tulu"),
    ("2028 A3", "Regulatsioon", "Täiskasvanute turu piloot; 15 puukooli; meditsiiniline tootmine; koolide looduskursused", "Must turg väheneb 30%", "280 töökohta, 12M tulu"),
    ("2029 A4", "Skaleerimine", "22 puukooli; 15 growshopi; eksport; päikesepargid fondist", "Riigisisene tarneahel", "450 töökohta, 22M tulu"),
    ("2030 A5", "Küpsus", "30 puukooli; 25 growshopi; masinad paid off → vaba kassa; basic income pilot 10k", "Eesmärk saavutatud", "650 töökohta, 35M tulu"),
]
for i, row in enumerate(plan, start=3):
    for j, val in enumerate(row, start=1):
        ws5.cell(row=i, column=j, value=val)
style_header(ws5, 3, 5)
style_table(ws5, 4, 8, 5)
ws5.row_dimensions[4].height = 50
ws5.row_dimensions[5].height = 50
ws5.row_dimensions[6].height = 50
ws5.row_dimensions[7].height = 50
ws5.row_dimensions[8].height = 50

# ─────────────────────────────────────────────
# LEHT 6: PUUKOOLID
# ─────────────────────────────────────────────
ws6 = wb.create_sheet("Puukoolid")
add_title(ws6, "PUUKOOLIDE JA KASVATUSETTEVÕTETE TOETUSMUDEL", 7)
set_widths(ws6, [22, 18, 14, 14, 14, 14, 35])

puu = [
    ("Toetusmeede", "Summa/üksus", "A1", "A2", "A3", "A4-A5", "Tingimused"),
    ("Seemikuprogramm (€/1000 tk)", 800, 3, 8, 15, 30, "Certified hemp seedlings"),
    ("Sertifitseerimisboonus (€/farm)", 5_000, 3, 8, 15, 30, "EU organic või equivalent"),
    ("Hektaritoetus (€/ha)", 350, 50, 200, 500, 1000, "Tööstuskanep, max 500€/ha"),
    ("Seadme liisingukaasfin (%)", "30%", 4, 12, 28, 60, "Kuivati, press, pakkimine"),
    ("Koolitustoetus (€/töötaja)", 500, 15, 40, 80, 150, "Kutsekool + EPM kursus"),
    ("Eksportboonus (€/tonn)", 120, 5, 20, 60, 120, "Kiud või seemned väljapoole"),
    ("Maaelu LEADER (olemasolev)", "Kuni 50k", "—", "—", "—", "—", "Taotlus läbi kohaliku KOV"),
    ("EAS arendusgrant (olemasolev)", "Kuni 200k", "—", "—", "—", "—", "Innovatsioon, tööstus"),
]
for i, row in enumerate(puu, start=3):
    for j, val in enumerate(row, start=1):
        ws6.cell(row=i, column=j, value=val)
style_header(ws6, 3, 7, fill=PatternFill("solid", fgColor="388E3C"))
style_table(ws6, 4, 11, 7)

# ─────────────────────────────────────────────
# LEHT 7: GROWSHOP VÕRDLUS
# ─────────────────────────────────────────────
ws7 = wb.create_sheet("Growshop mudel")
add_title(ws7, "GROWSHOP — VÄLISMAA VS EESTI RIIKLIK MUDEL", 6)
set_widths(ws7, [24, 20, 20, 20, 20, 35])

grow = [
    ("Näitaja", "Välismaa growshop", "Eesti Roheline", "Eelis", "A5 eesmärk", "Märkus"),
    ("Omanik", "Hollandi/Saksa kett", "ERMA 51% + kohalik", "Raha Eestis", "25 poodi", "Vaba turg, parem teenus"),
    ("Hind 1.33g", "15 €", "12-14 €", "Odavam", "10-12 €", "Otsetootja, vahendajad välja"),
    ("Kvaliteedikontroll", "Puudub/ebaühtlane", "Riiklik labor", "Ohutus", "100% testitud", "Partii kaupa"),
    ("Klienditeenindus", "Online, anonüümne", "Eesti keeles, nõustamine", "Inimlik", "NPS > 70", "Meditsiiniline info"),
    ("Maksud Eestisse", "0 €", "Käibemaks + aktsiis", "Riigile", "8M €/a tulu", "Must turg → legal"),
    ("Töökohad/pood", "1-2", "4-6", "Kohalik", "150+ retail", "Hooldus, nõustamine"),
    ("Toote päritolu", "Import", "70% kohalik A5", "Suveräänsus", "90% kohalik", "Eesti kanep"),
]
for i, row in enumerate(grow, start=3):
    for j, val in enumerate(row, start=1):
        ws7.cell(row=i, column=j, value=val)
style_header(ws7, 3, 6)
style_table(ws7, 4, 10, 6)

# ─────────────────────────────────────────────
# LEHT 8: LIISEINGUMUDEL
# ─────────────────────────────────────────────
ws8 = wb.create_sheet("Liisingumudel")
add_title(ws8, "MASINAD LIISEINGUGA — VABA KASSA MUDEL", 7)
set_widths(ws8, [24, 14, 14, 14, 14, 14, 38])

lease = [
    ("Seade", "Hind (€)", "Liising (kuu)", "Kestus", "Klient maksab", "Pärast lõppu", "Selgitus"),
    ("Kuivati 500kg", 85_000, 1_850, "48 kk", "89 €/kuu lisatasu", "ERMA omand", "Farmer lease → paid off"),
    ("Õlipress", 120_000, 2_600, "48 kk", "125 €/kuu", "ERMA omand", "CBD kosmeetika liin"),
    ("Pakkeliin", 65_000, 1_420, "48 kk", "68 €/kuu", "ERMA omand", "Growshop supply"),
    ("Labor seade", 45_000, 980, "48 kk", "—", "ERMA omand", "Kvaliteedikontroll"),
    ("Kasvuruumi tech", 35_000, 760, "48 kk", "40 €/kuu", "ERMA/partner", "Puukooli partner"),
    ("", "", "", "", "", "", ""),
    ("VABA KASSA loogika:", "", "", "", "", "", ""),
    ("Aastad 1-4", "Klient maksab toote + liisinguosa", "", "", "", "", "Masin ostetakse välja"),
    ("Aasta 5+", "Masin on riigi/partneri oma", "", "", "", "", "Kasum ilma liisinguta = vaba kassa"),
    ("5a liisingutulu", 400_000, 1_000_000, 2_000_000, 3_000_000, 3_500_000, "Kumulatiivne reinvesteering"),
]
for i, row in enumerate(lease, start=3):
    for j, val in enumerate(row, start=1):
        ws8.cell(row=i, column=j, value=val)
style_header(ws8, 3, 7, fill=PatternFill("solid", fgColor="F9A825"))
style_table(ws8, 4, 13, 7)
for r in range(4, 9):
    ws8.cell(row=r, column=2).number_format = '#,##0'
    ws8.cell(row=r, column=3).number_format = '#,##0'

# ─────────────────────────────────────────────
# LEHT 9: TÖÖKOHALD
# ─────────────────────────────────────────────
ws9 = wb.create_sheet("Töökohad")
add_title(ws9, "TÖÖKOHTADE PROGNOOS", 6)
set_widths(ws9, [26, 12, 12, 12, 12, 12, 30])

jobs = [
    ("Valdkond", "A1", "A2", "A3", "A4", "A5", "Keskmine palk (bruto)"),
    ("Puukoolid / kasvatus", 12, 35, 80, 120, 160, "1 400 €"),
    ("Töötlemine / tootmine", 8, 25, 60, 100, 150, "1 800 €"),
    ("Growshop / jaemüük", 5, 25, 60, 100, 150, "1 500 €"),
    ("Labor / kvaliteet", 3, 8, 15, 20, 25, "2 200 €"),
    ("IT / logistika", 4, 10, 20, 35, 50, "2 000 €"),
    ("Juhtimine / haldus", 5, 8, 15, 25, 35, "2 500 €"),
    ("Koolitus / nõustamine", 3, 5, 12, 20, 30, "1 700 €"),
    ("Ehitus / hampcrete", 5, 4, 18, 30, 50, "1 600 €"),
    ("KOKKU", "=SUM(B2:B9)", "=SUM(C2:C9)", "=SUM(D2:D9)", "=SUM(E2:E9)", "=SUM(F2:F9)", ""),
    ("", "", "", "", "", "", ""),
    ("Palgafond (hinnanguline €)", 900_000, 1_400_000, 2_500_000, 3_800_000, 5_200_000, "Sh sotsiaalmaks"),
]
for i, row in enumerate(jobs, start=3):
    for j, val in enumerate(row, start=1):
        ws9.cell(row=i, column=j, value=val)
style_header(ws9, 3, 7)
style_table(ws9, 4, 13, 7)

# ─────────────────────────────────────────────
# LEHT 10: BRÜSSELL
# ─────────────────────────────────────────────
ws10 = wb.create_sheet("Brüssel")
add_title(ws10, "MIDA BRÜSSELILE ÖELDA — INIMLIK POSITSIOON", 3)
set_widths(ws10, [22, 50, 50])

bru = [
    ("Teema", "Meie sõnum Brüsselile", "Õiguslik alus / precedents"),
    ("Tööstuskanep", "Eesti soovib olla EL tööstuskanepi keskus. Kiud, ehitus, toit — kõik juba lubatud.", "EL Common Agricultural Policy; hemp <0.3% THC"),
    ("CBD kosmeetika", "Eesti toodab kosmeetikat, mitte toitu. EFSA novel food ei kehti kreemidele.", "Cosmetics Regulation 1223/2009"),
    ("Siseriiklik reguleerimine", "Täiskasvanute kanep on siseriiklik küsimus. Saksamaa, Malta, Holland on eeskujuks.", "Subsidiarity; EU no harmonized THC policy"),
    ("Rural development", "Kanep aitab maapiirkondi. Palume CAP Pillar II toetusi kasvatamiseks.", "EAFRD, LEADER, Green Deal"),
    ("Harm reduction", "Reguleerimine vähendab musta turgu ja noorte juurdepääsu. Inimkeskne lähenemine.", "WHO; German Cannabis Act rationale"),
    ("Kliimaeesmärgid", "Kanep asendab plastikut, sidub CO₂, toetab rohelist üleminekut.", "EU Green Deal; Fit for 55"),
    ("Mida ME EI ütle", "Me ei küsi luba 'narkootikumide müümiseks'. Me teavitame tööstus- ja maaeluplaanist.", "—"),
    ("Mida ME KÜSIME", "Toetust tööstuskanepi töötlemiseks; selgust CBD kosmeetikas; subsidiarity respect.", "Horizon; EIC; CAP"),
]
for i, row in enumerate(bru, start=3):
    for j, val in enumerate(row, start=1):
        ws10.cell(row=i, column=j, value=val)
style_header(ws10, 3, 3, fill=BLUE.fill_type and PatternFill("solid", fgColor="1565C0") or BLUE)
style_table(ws10, 4, 11, 3)
for r in range(4, 12):
    ws10.row_dimensions[r].height = 45

# ─────────────────────────────────────────────
# LEHT 11: OLEMASOLEVAD VAHENDID
# ─────────────────────────────────────────────
ws11 = wb.create_sheet("Olemasolevad vahendid")
add_title(ws11, "OLEMASOLEVAD VAHENDID — INIMESELT ILMA UUT RAHA TAOTLEMATA", 5)
set_widths(ws11, [24, 18, 18, 18, 45])

vahendid = [
    ("Vahend", "Summa", "Kellele", "Kuidas kasutada", "Inimlik samm"),
    ("LEADER / maaelu", "Kuni 50 000 €", "Puukool, väikefarm", "Seemikprogramm, kuivati", "KOV kaasab, inimesed ei jää üksi"),
    ("EAS arendusgrant", "Kuni 200 000 €", "ERMA, startup", "Growshop võrgustik, IT", "Eesti ettevõte, mitte Holland"),
    ("KIK kliimafond", "Projektipõhine", "Töötlemine", "Päikesepark tööstuse juurde", "Roheline energia → odavam tootmine"),
    ("Horizon Europe", "Miljonid (konsortsium)", "Teadus + EPM", "Kanepi uuringud, sordid", "Teaduslik usaldus Brüsselile"),
    ("Kutsekoolide rahastus", "Olemasolev eelarve", "Kutsekoolid", "Uus eriala: taimekasvatus", "Lapsed näevad karjääri"),
    ("Linnaplaneerimine", "KOV eelarve", "Koolid, linnad", "Kooli aiad, urban farm", "Loodus tagasi linnadesse"),
    ("Riigihanke raam", "Olemasolev", "Haiglad, hooldekodud", "CBD kosmeetika, hampcrete", "Kohalik toode avalikus sektoris"),
    ("Töötukassa koolitus", "Töötute programm", "Ümberõpe", "Kanepitööstuse kursused", "Inimene saab uue oskuse"),
]
for i, row in enumerate(vahendid, start=3):
    for j, val in enumerate(row, start=1):
        ws11.cell(row=i, column=j, value=val)
style_header(ws11, 3, 5)
style_table(ws11, 4, 11, 5)
for r in range(4, 12):
    ws11.row_dimensions[r].height = 40

# ─────────────────────────────────────────────
# LEHT 12: KOOLID JA LOODUS
# ─────────────────────────────────────────────
ws12 = wb.create_sheet("Koolid ja loodus")
add_title(ws12, "LINNAPLANEERIMINE JA KOOLID — UUED ELUALAD", 4)
set_widths(ws12, [20, 35, 35, 40])

kool = [
    ("Tase", "Mis muutub", "Konkreetne tegevus", "Tulemus 5 aasta pärast"),
    ("Põhikool", "Looduskursus", "Kooli aed + kanepi demo (tööstus)", "Iga laps teab, mis on taimetööstus"),
    ("Gümnaasium", "Karjääripäev", "EPM, farmid, labor külastus", "30% rohkem taimekasvatus eriala"),
    ("Kutsekool", "Uus eriala", "Taimekasvatus ja töötlemine", "200 lõpetajat/aasta tööturule"),
    ("Ülikool", "Tehnoloogia", "Biokomposiidid, CBD teadus", "Eesti patendid, eksport"),
    ("Linn", "Urban farm", "Katusaared, kogukonna aiad", "Loodus 500m kaugusel igast kodust"),
    ("KOV", "Planeering", "Tööstuskanep tsoonid, rohekoridorid", "Tööstus + elu kooskõlas"),
]
for i, row in enumerate(kool, start=3):
    for j, val in enumerate(row, start=1):
        ws12.cell(row=i, column=j, value=val)
style_header(ws12, 3, 4)
style_table(ws12, 4, 9, 4)
for r in range(4, 10):
    ws12.row_dimensions[r].height = 40

# ─────────────────────────────────────────────
# LEHT 13: FINANTSSTRATEEGIA — DALIO / ROBBINS
# ─────────────────────────────────────────────
ws13 = wb.create_sheet("Finantsstrateegia")
add_title(ws13, "EESTI FINANTSSTRATEEGIA — 18 MITTE-KORRELATSIOONSET VOOGU (RAY DALIO / TONY ROBBINS)", 5)
set_widths(ws13, [8, 30, 22, 12, 40])

ws13.merge_cells("A2:E2")
ws13.cell(
    row=2,
    column=1,
    value=(
        "Ray Dalio põhimõte (Tony Robbins *All Seasons*): leia 15–20 mitte-korrelatsioonset tuluvoogu. "
        "ERMA kasum → Eesti Tuleviku Fond jaotatakse nii, et üks kriis ei võtab kõike."
    ),
).alignment = Alignment(wrap_text=True)

dalio = [
    ("#", "Tuluvoog / varaklass", "Korrelatsioon", "Siht %", "Märkus"),
    ("1", "Päikesepargid + salvestus", "Energia ≠ tarbimine", 0.12, "Macro-kott 40% alamvoog"),
    ("2", "Liisingumasinad", "≠ börs", 0.08, "Vaba kassa A5+"),
    ("3", "Tööstuskanep eksport", "Ekspordi tsükkel", 0.07, "Kiud, seemned"),
    ("4", "CBD kosmeetika", "Tarbimine", 0.05, "Kreemid, õlid"),
    ("5", "Growshop võrgustik", "Kohalik müük", 0.06, "25 poodi A5"),
    ("6", "Reguleeritud täiskasvanute turg", "Siseriiklik", 0.05, "A3 piloot"),
    ("7", "Meditsiiniline kanep", "Haigekassa", 0.04, "Stabiilne ostja"),
    ("8", "Puukoolid + seemik", "Maaelu", 0.06, "30 puukooli"),
    ("9", "Kutsekool + koolitus", "Inimkapital", 0.05, "Uus eriala"),
    ("10", "Operation Mirror säästud", "Bürokraatia", 0.05, "Koostöö Mirror mudeliga"),
    ("11", "LEADER / EAS / CAP", "Grantid", 0.04, "Olemasolev raha"),
    ("12", "Urban farm + KOV", "Kogukond", 0.03, "Linn + maa"),
    ("13", "Teadus + patendid", "Pikaajaline", 0.04, "EPM"),
    ("14", "Hampcrete + ehitus", "Ehitus", 0.03, "Riigihanke"),
    ("15", "Kriisitoit / varud", "Defensiivne", 0.03, "2-nädala reegel"),
    ("16", "Turism + kogemus", "Hooajaline", 0.02, "Farm visits"),
    ("17", "IT + IoT kasvatus", "Tehnoloogia", 0.03, "Efektiivsus"),
    ("18", "Basic income pilot", "Sotsiaal", 0.06, "Macro-kott 20%"),
    ("", "KOKKU", "18 voogu (Dalio: 15–20)", "=SUM(D4:D21)", "100% fondi"),
]
for i, row in enumerate(dalio, start=4):
    for j, val in enumerate(row, start=1):
        ws13.cell(row=i, column=j, value=val)
style_header(ws13, 4, 5, fill=PatternFill("solid", fgColor="6A1B9A"))
style_table(ws13, 5, 22, 5)
for r in range(5, 22):
    ws13.cell(row=r, column=4).number_format = "0.0%"

wb.save(OUTPUT)
print(f"Salvestatud: {OUTPUT}")
