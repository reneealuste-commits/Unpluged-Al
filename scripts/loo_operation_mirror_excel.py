#!/usr/bin/env python3
"""Operation Mirror + Hemp Authority pivot — Excel finantsmudel."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/workspace/Operation-Mirror-Pivot-Finants.xlsx"

GREEN = PatternFill("solid", fgColor="2E7D32")
DARK = PatternFill("solid", fgColor="1B5E20")
GOLD = PatternFill("solid", fgColor="FFF8E1")
BLUE = PatternFill("solid", fgColor="E3F2FD")
RED = PatternFill("solid", fgColor="FFEBEE")
GRAY = PatternFill("solid", fgColor="F5F5F5")
LIGHT_GREEN = PatternFill("solid", fgColor="E8F5E9")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1B5E20")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EUR = '#,##0 "€"'
PCT = '0.0%'


def style_header(ws, row, cols, fill=GREEN):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_table(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r % 2 == 0:
                cell.fill = GRAY


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_title(ws, title, merge_cols=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = LIGHT_GREEN
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


def write_rows(ws, start_row, rows):
    for i, row in enumerate(rows):
        for j, val in enumerate(row, 1):
            ws.cell(row=start_row + i, column=j, value=val)
    return start_row + len(rows)


def fmt_eur(ws, cells):
    for ref in cells:
        ws[ref].number_format = EUR


wb = Workbook()

# ═══════════════════════════════════════════════════════════
# LEHT 1: KOKKUVÕTE
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Kokkuvõte"
add_title(ws, "OPERATION MIRROR + HEMP AUTHORITY — RAHA KOKKUVÕTE", 7)
set_widths(ws, [32, 16, 16, 16, 16, 16, 38])

summary = [
    ("", "Aasta 1", "Aasta 3", "Aasta 5", "5a KOKKU", "Märkus", ""),
    ("RAHA SISSE", "", "", "", "", "", ""),
    ("Musta turu maksud (reguleeritud)", 0, 18_000_000, 46_000_000, "=SUM(B4:D4)", "Julgeolekuteaduste uuring: 46M/a", ""),
    ("ERMA kasum riigikassasse", 0, 1_800_000, 7_500_000, "=SUM(B5:D5)", "Eesti Roheline Majandus AS", ""),
    ("Operation Mirror säästud", 5_000_000, 25_000_000, 35_000_000, "=SUM(B6:D6)", "Bürokraatia kärpimine", ""),
    ("Tööstuskanep + eksport (maksud)", 3_000_000, 8_000_000, 15_000_000, "=SUM(B7:D7)", "Nordic Hemp jms", ""),
    ("EL vahendid (LEADER, EAS, CAP)", 2_000_000, 3_000_000, 5_000_000, "=SUM(B8:D8)", "Olemasolev, mitte uus maks", ""),
    ("KOKKU SISSE", "=SUM(B4:B8)", "=SUM(C4:C8)", "=SUM(D4:D8)", "=SUM(E4:E8)", "", ""),
    ("", "", "", "", "", "", ""),
    ("RAHA VÄLJA", "", "", "", "", "", ""),
    ("Hemp Authority admin", 3_000_000, 4_000_000, 5_000_000, "=SUM(B12:D12)", "80→150 töötajat", ""),
    ("ERMA tegevuskulud", 2_800_000, 6_500_000, 12_000_000, "=SUM(B13:D13)", "Tootmine, palgad, liising", ""),
    ("Puukoolide toetused", 300_000, 600_000, 400_000, "=SUM(B14:D14)", "Boonused farmeritele", ""),
    ("Sotsiaal/tervise kulud (regulatsioon)", 2_000_000, 6_000_000, 10_000_000, "=SUM(B15:D15)", "Julgeolekuteaduste hinnang", ""),
    ("KOKKU VÄLJA", "=SUM(B12:B15)", "=SUM(C12:C15)", "=SUM(D12:D15)", "=SUM(E12:E15)", "", ""),
    ("", "", "", "", "", "", ""),
    ("NETO INIMESTELE (Inimeste Fond)", "=B9-B16", "=C9-C16", "=D9-D16", "=SUM(B18:D18)", "Jaotatakse rahvale", ""),
    ("Neto €/elanik (1,37M)", "=B18/1370000", "=C18/1370000", "=D18/1370000", "", "Aasta 5: ~68€/inimene", ""),
    ("", "", "", "", "", "", ""),
    ("PÕHIMÕTE", "Operation Mirror", "", "", "", "Peegeldab bürokraatiat, ei hävita taristut", ""),
    ("", "Hemp Authority", "", "", "", "Maksab enda eest A3-st", ""),
    ("", "Inimeste Fond", "", "", "", "Dalio: 8 voogu; macro 40/30/20/10", ""),
]
write_rows(ws, 3, summary)
style_header(ws, 3, 7)
style_table(ws, 4, 24, 7)
for r in range(4, 20):
    for c in range(2, 6):
        ws.cell(row=r, column=c).number_format = EUR
for r in [19]:
    for c in range(2, 5):
        ws.cell(row=r, column=c).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 2: KLIIMAMINISTEERIUM — MIDA SAAB / EI SAA
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Kliimaministeerium")
add_title(ws2, "KLIIMAMINISTEERIUM 2025 — MIDA SAAB PÖÖRATA", 6)
set_widths(ws2, [34, 18, 14, 14, 40])

climate = [
    ("Programm / kulu", "Eelarve 2025", "Pivot?", "Realistlik sääst", "Selgitus"),
    ("KOKKU valitsemisala", 1_061_402_000, "—", "—", "Riigieelarve andmed"),
    ("Transpordi programm (Rail Baltica, teed)", 643_949_000, "EI", 0, "Taristu — ei puutu"),
    ("Elukeskkond + ehitus", 162_117_000, "OSALISELT", 8_000_000, "Streamline, mitte kustuta"),
    ("Kliima + energia + elurikkus", 171_763_000, "JAH", 15_000_000, "Pivot zone — kanep asendab osa"),
    ("Mere + vesi", 63_909_000, "EI", 0, "Essential services"),
    ("Ministeeriumi admin (~308 töötajat)", 20_000_000, "JAH", 12_000_000, "Operation Mirror sihtmärk"),
    ("Rohereformi bürokraatia", 34_746_000, "JAH", 8_000_000, "PDF → taimed"),
    ("", "", "", "", ""),
    ("REALISTLIK KOKKU SÄÄST", "", "", 35_000_000, "Aasta 5 eesmärk"),
    ("MIDA EI PUUTU", "", "", 707_858_000, "Transport + vesi + meri"),
]
write_rows(ws2, 3, climate)
style_header(ws2, 3, 5)
style_table(ws2, 4, 14, 5)
for r in range(4, 14):
    ws2.cell(row=r, column=2).number_format = EUR
    if ws2.cell(row=r, column=4).value and isinstance(ws2.cell(row=r, column=4).value, (int, float)):
        ws2.cell(row=r, column=4).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 3: 4 RAHAALLIKAT
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("4 rahaallikat")
add_title(ws3, "4 RAHAALLIKAT — KUST RAHA TULEB", 6)
set_widths(ws3, [28, 18, 18, 18, 18, 36])

sources = [
    ("Allikas", "A1", "A3", "A5", "Kumulatiivne", "Kuidas"),
    ("1. MUST TURG → REGULEERITUD", "", "", "", "", ""),
    ("Tarbijate kulutus (must turg)", 56_000_000, 80_000_000, 120_000_000, "=SUM(B3:D3)", "56M uuring / 172M petitsioon"),
    ("Riigile (maksud + säästud)", 0, 18_000_000, 46_000_000, "=SUM(B4:D4)", "15% aktsiis + KM + politsei"),
    ("", "", "", "", "", ""),
    ("2. OPERATION MIRROR", "", "", "", "", ""),
    ("Bürokraatia säästud", 5_000_000, 20_000_000, 35_000_000, "=SUM(B7:D7)", "Ministeeriumide peegeldamine"),
    ("", "", "", "", "", ""),
    ("3. ERMA RIIGIETTEVÕTE", "", "", "", "", ""),
    ("ERMA tulud", 1_200_000, 12_000_000, 35_000_000, "=SUM(B10:D10)", "Growshop, tööstus, liising"),
    ("ERMA kasum", -1_600_000, 5_500_000, 23_000_000, "=SUM(B11:D11)", "Vaba kassa A5+"),
    ("Riigikassasse", 0, 1_800_000, 7_500_000, "=SUM(B12:D12)", "Maksud + dividendid"),
    ("", "", "", "", "", ""),
    ("4. OLEMASOLEVAD VAHENDID", "", "", "", "", ""),
    ("LEADER + EAS + CAP + KIK", 2_000_000, 3_000_000, 5_000_000, "=SUM(B15:D15)", "Ilma uue maksuta"),
    ("", "", "", "", "", ""),
    ("KOKKU RIIGILE A5", "=B4+B7+B12+B15", "=C4+C7+C12+C15", "=D4+D7+D12+D15", "=SUM(B17:D17)", ""),
]
write_rows(ws3, 3, sources)
style_header(ws3, 3, 6)
style_table(ws3, 4, 18, 6)
for r in [3, 4, 7, 10, 11, 12, 15, 17]:
    for c in range(2, 6):
        if ws3.cell(row=r, column=c).value != "":
            ws3.cell(row=r, column=c).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 4: VANA VS UUS MUDEL
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Vana vs Uus")
add_title(ws4, "VANA MUDEL vs HEMP AUTHORITY", 5)
set_widths(ws4, [28, 22, 22, 22, 30])

compare = [
    ("Näitaja", "Kliimaministeerium (lõik)", "Hemp Authority + ERMA", "Erinevus A5", "Märkus"),
    ("Töötajad", 308, 150, -158, "Lean team"),
    ("Admin kulu/a", 20_000_000, 5_000_000, -15_000_000, "80% vähem bürokraatiat"),
    ("Tulu riigile/a", 0, 53_000_000, 53_000_000, "46M maksud + 7.5M ERMA"),
    ("Neto rahvale/a", -20_000_000, 48_000_000, 68_000_000, "Maksab enda + jagab"),
    ("Töökohti loodud", 0, 650, 650, "ERMA võrgustik"),
    ("Must turg", "100% must", "30% must A5", "—", "Reguleerimine vähendab"),
    ("Päikesepargid", "40M invest (riik maksab)", "28M fondist (tulu maksab)", "—", "Self-funding"),
    ("Basic income pilot", 0, "10 000 × 200€/kuu", "24M/a", "Inimeste Fondist"),
]
write_rows(ws4, 3, compare)
style_header(ws4, 3, 5)
style_table(ws4, 4, 12, 5)
for r in range(4, 12):
    for c in [2, 3, 4]:
        v = ws4.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            ws4.cell(row=r, column=c).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 5: AASTA 5 TÄIELIK PILT
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Aasta 5 pilt")
add_title(ws5, "AASTA 5 (2030) — TÄIELIK RAHA VOOG", 4)
set_widths(ws5, [36, 18, 18, 40])

y5 = [
    ("", "Summa", "%", "Selgitus"),
    ("RAHA SISSE", "", "", ""),
    ("Reguleeritud turu maksud", 46_000_000, 0.37, "KM + aktsiis + tööjõumaksud"),
    ("ERMA kasum + dividendid", 7_500_000, 0.06, "Riigiettevõte"),
    ("Operation Mirror säästud", 35_000_000, 0.28, "Bürokraatia"),
    ("Tööstuskanep maksud", 15_000_000, 0.12, "Eksport + kohalik"),
    ("EL vahendid", 5_000_000, 0.04, "LEADER, CAP jne"),
    ("Muu (turism, koolitus)", 3_000_000, 0.02, ""),
    ("KOKKU SISSE", "=SUM(B3:B8)", 1.0, "≈ 111M €"),
    ("", "", "", ""),
    ("RAHA VÄLJA", "", "", ""),
    ("Hemp Authority admin", 5_000_000, 0.05, ""),
    ("ERMA tegevuskulud", 12_000_000, 0.11, ""),
    ("Puukoolide toetused", 400_000, 0.00, ""),
    ("Sotsiaal/tervise kulud", 10_000_000, 0.09, ""),
    ("KOKKU VÄLJA", "=SUM(B11:B14)", 0.25, "≈ 27M €"),
    ("", "", "", ""),
    ("NETO INIMESTELE", "=B9-B15", 0.75, "≈ 84M €"),
]
write_rows(ws5, 3, y5)
style_header(ws5, 3, 4)
style_table(ws5, 4, 19, 4)
for r in range(3, 20):
    if r != 9 and r != 15:
        ws5.cell(row=r, column=2).number_format = EUR
    if ws5.cell(row=r, column=3).value and isinstance(ws5.cell(row=r, column=3).value, float):
        ws5.cell(row=r, column=3).number_format = PCT

# ═══════════════════════════════════════════════════════════
# LEHT 6: INIMESTE FOND
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Inimeste Fond")
add_title(ws6, "INIMESTE FOND — KUHU NETO LÄHEB (A5)", 5)
set_widths(ws6, [30, 14, 14, 14, 32])

fund = [
    ("Kategooria", "%", "Summa A5", "Mida ostab", "Näide"),
    ("Päikesepargid + salvestus", 0.40, "=B3*$B$9", "8× 5MW parki", "28M € — odavam elekter"),
    ("Tööstus + robotid", 0.30, "=B4*$B$9", "Liisingumasinad", "1000 robotit / automatiseerimine"),
    ("Basic income pilot", 0.20, "=B5*$B$9", "10 000 inimest", "200 €/kuu × 12 = 24M"),
    ("Haridus + koolid + loodus", 0.10, "=B6*$B$9", "Kooli aiad, karjäär", "Urban farm, kutsekool"),
    ("", "", "", "", ""),
    ("NETO FONDI SUURUS A5", "", 84_000_000, "", "Eelarve real B9 lehel Aasta 5"),
    ("", "", "", "", ""),
    ("Per capita (1,37M elanikku)", "", "=B9/1370000", "€/aasta/inimene", ""),
    ("BI pilot inimeste arv", 10_000, "", "200€/kuu", "24M/a"),
    ("Töökohad loodud", 650, "", "ERMA võrgustik", ""),
    ("Päikesevõimsus MW", 40, "", "8 parki × 5MW", ""),
]
write_rows(ws6, 3, fund)
ws6.cell(row=9, column=3, value="=B9")  # reference placeholder - fix with cross-sheet
# Use fixed value for fund size since cross-sheet is complex
ws6["C4"] = "=C9*0.4"
ws6["C5"] = "=C9*0.3"
ws6["C6"] = "=C9*0.2"
ws6["C7"] = "=C9*0.1"
style_header(ws6, 3, 5, fill=PatternFill("solid", fgColor="F9A825"))
style_table(ws6, 4, 13, 5)
ws6["B4"].number_format = PCT
ws6["B5"].number_format = PCT
ws6["B6"].number_format = PCT
ws6["B7"].number_format = PCT
for ref in ["C4", "C5", "C6", "C7", "C9", "C11"]:
    ws6[ref].number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 7: 5-SAMMULINE PIVOT
# ═══════════════════════════════════════════════════════════
ws7 = wb.create_sheet("5-sammuline pivot")
add_title(ws7, "5-SAMMULINE PIVOT — KUIDAS TEHA", 7)
set_widths(ws7, [10, 22, 16, 16, 16, 16, 34])

pivot = [
    ("Samm", "Aeg", "Vaja raha", "Kust tuleb", "Tulemus", "Risk", "Tegevus"),
    ("1", "Kohe", 0, "Tarbijaliikumine", "Nõudlus nähtav", "Madal", "Osta Eesti, räägi 3 sõbrale"),
    ("2", "6 kuud", 500_000, "EAS 200k + LEADER 150k + partner 150k", "ERMA piloot", "Madal", "3 puukooli, bränd"),
    ("3", "A1 (2026)", 2_800_000, "EL vahendid + laen + ERMA", "45 töökohta", "Keskmine", "5 growshopi, klaster"),
    ("4", "A2-A3", 10_000_000, "Mirror säästud + ERMA kasum", "280 töökohta", "Keskmine", "Hemp Authority loomine"),
    ("5", "A4-A5", 0, "Isetoetav", "650 tk, 84M fond", "Madal", "Reguleeritud turg, BI pilot"),
    ("", "", "", "", "", "", ""),
    ("KOKKU käivituskapital", "", 3_300_000, "Mitte ministeeriumi eelarvest!", "", "", "A1+A2 bootstrap"),
    ("A5 isetoetav?", "", "JAH", "53M sisse > 27M välja", "", "", "18× admin kulu"),
]
write_rows(ws7, 3, pivot)
style_header(ws7, 3, 7)
style_table(ws7, 4, 11, 7)
for r in [5, 6, 7, 8, 10]:
    ws7.cell(row=r, column=3).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 8: OPERATION MIRROR DETAIL
# ═══════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Operation Mirror")
add_title(ws8, "OPERATION MIRROR — MIDA PEEGELDADA", 5)
set_widths(ws8, [32, 18, 18, 18, 36])

mirror = [
    ("Funktsioon", "Praegu", "Mirror tegevus", "Sääst/a", "Tulemus"),
    ("Kliimapoliitika raportid", "35M programm", "Tulemuspõhine ainult", 5_000_000, "PDF → taimed"),
    ("Agri + kliima kanep kattuvus", "2 ministeeriumi", "Üks Hemp Authority", 3_000_000, "Üks luba, üks lab"),
    ("6 eraldi IT süsteemi", "16M IT kulud", "Üks platvorm", 4_000_000, "PRIA + ERMA integreeritud"),
    ("Konsultandid + komisjonid", "~8M/a", "Farmerite kooperatiiv otsustab", 5_000_000, "Demokraatia maal"),
    ("Välismaa CBD import", "15-30M välja", "ERMA asendab", 15_000_000, "Raha Eestis"),
    ("Ministeeriumide arv 11→7", "~40M admin", "Streamline 4 ministeeriumi", 8_000_000, "Operation Mirror täis"),
    ("", "", "KOKKU A5", 35_000_000, "Realistlik eesmärk"),
    ("", "", "EI PUUTU", 707_858_000, "Teed, Rail Baltica, vesi"),
]
write_rows(ws8, 3, mirror)
style_header(ws8, 3, 5, fill=DARK)
style_table(ws8, 4, 11, 5)
for r in range(4, 11):
    for c in [2, 4]:
        v = ws8.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            ws8.cell(row=r, column=c).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 9: BOOTSTRAP — VÄIKE RAHA KOHE
# ═══════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Bootstrap raha")
add_title(ws9, "KUI VAJAD VÄHESELT RAHA — KOHE", 5)
set_widths(ws9, [24, 18, 18, 18, 40])

boot = [
    ("Vajad", "Summa", "Allikas", "Aeg", "Kuidas taotleda"),
    ("1 puukool", 50_000, "LEADER", "3-6 kuud", "KOV kaudu, projektikiri"),
    ("IT + bränd", 200_000, "EAS arendusgrant", "4-8 kuud", "eas.ee, ettevõte vajalik"),
    ("Seemikuprogramm", 30_000, "CAP Pillar II", "6 kuud", "Põllumajandusministeerium"),
    ("Labor seade", 45_000, "KIK kliimafond", "6-12 kuud", "Kvaliteedikontroll"),
    ("Päikesepaneelid tööstusele", 80_000, "KIK", "6-12 kuud", "Odavam tootmine"),
    ("Kutsekooli programm", 0, "Olemasolev eelarve", "12 kuud", "Uus eriala, ei vaja uut raha"),
    ("", "", "", "", ""),
    ("KOKKU bootstrap", 405_000, "Ilma ministeeriumita!", "", ""),
    ("+ Nordic Hemp partner", 150_000, "Ettevõte", "Kohe", "Kasvatajate võrgustik"),
    ("KOKKU piloot", 555_000, "", "6 kuud", "ERMA käivitamiseks piisav"),
]
write_rows(ws9, 3, boot)
style_header(ws9, 3, 5, fill=BLUE)
style_table(ws9, 4, 13, 5)
for r in range(4, 13):
    v = ws9.cell(row=r, column=2).value
    if isinstance(v, (int, float)):
        ws9.cell(row=r, column=2).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 10: ERMA 5-AASTA (viide)
# ═══════════════════════════════════════════════════════════
ws10 = wb.create_sheet("ERMA 5a")
add_title(ws10, "ERMA AS — 5-AASTA NUMBRID", 6)
set_widths(ws10, [24, 14, 14, 14, 14, 14])

erma = [
    ("", "A1", "A2", "A3", "A4", "A5"),
    ("Tulud", 1_200_000, 4_500_000, 12_000_000, 22_000_000, 35_000_000),
    ("Kulud", 2_800_000, 3_600_000, 6_500_000, 9_000_000, 12_000_000),
    ("Netokasum", "=B3-B4", "=C3-C4", "=D3-D4", "=E3-E4", "=F3-F4"),
    ("Riigikassasse", 0, 350_000, 1_800_000, 4_200_000, 7_500_000),
    ("Töökohad", 45, 120, 280, 450, 650),
    ("Liisingumasinad", 4, 12, 28, 45, 60),
    ("Puukoolid", 3, 8, 15, 22, 30),
    ("Vaba kassa masinad", 0, 2, 8, 20, 45),
]
write_rows(ws10, 3, erma)
style_header(ws10, 3, 6)
style_table(ws10, 4, 11, 6)
for r in range(4, 8):
    for c in range(2, 7):
        ws10.cell(row=r, column=c).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 11: MUST TURG vs REGULEERITUD
# ═══════════════════════════════════════════════════════════
ws11 = wb.create_sheet("Must turg vs legal")
add_title(ws11, "MUST TURG vs REGULEERITUD — SINU HINNAGA", 6)
set_widths(ws11, [28, 16, 16, 16, 16, 30])

market = [
    ("Stsenaarium", "Turu maht", "Riigile", "Töökohti", "Must %", "Allikas"),
    ("Praegu (must turg)", 56_000_000, 0, 0, 1.0, "Julgeolekuteaduste uuring"),
    ("Petitsioon hinnang", 171_500_000, 0, 0, 1.0, "O. Lesment"),
    ("", "", "", "", "", ""),
    ("A3 piloot (30% legal)", 56_000_000, 18_000_000, 280, 0.7, "Meie mudel"),
    ("A5 täismahus", 80_000_000, 46_000_000, 650, 0.3, "Meie mudel"),
    ("", "", "", "", "", ""),
    ("Sinu hind", "15€/1.33g", "=15/1.33", "€/g", 11.28, ""),
    ("25k ostja × 2g/kuu", 6_770_000, 1_354_000, "", "", "Ainult see grupp"),
    ("Kogu Eesti (A5)", 80_000_000, 46_000_000, "", "", "Täismahus"),
]
write_rows(ws11, 3, market)
style_header(ws11, 3, 6)
style_table(ws11, 4, 11, 6)
for r in [4, 5, 8, 9, 10, 11]:
    for c in range(2, 5):
        v = ws11.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            ws11.cell(row=r, column=c).number_format = EUR

# ═══════════════════════════════════════════════════════════
# LEHT 12: ÕLESSANDED ALLÜKSUSTELE
# ═══════════════════════════════════════════════════════════
ws12 = wb.create_sheet("Olessanded")
add_title(ws12, "ÕLESSANDED ALLÜKSUSTELE — MALEVAPEALIKUD MAASTIKUL", 6)
set_widths(ws12, [22, 28, 28, 28, 22, 30])

oless = [
    ("Kiht", "Kes", "Mida teeb maastikul", "Mõju elanikele", "A5 eesmärk", "Seos teiste kihtidega"),
    ("7. OLESSANDED", "Malevapealik", "Koordineerib tsiviiltegevust", "Töökohad, turvalisus", "15 maleva aktiivne", "Operation Mirror elluviija"),
    ("Õlg", "Malevkonna pealik", "Kuulab vallas/linnas", "Farmer↔KOV↔tarbija", "150+ kuulamispunkti", "ERMA nõudlus"),
    ("Õlg", "JUP 2 lõpetaja", "Probleemide lahendaja", "Kriisikomisjonid", "500+ juhti ühiskonnas", "KOV Šveitsi mudel"),
    ("Õlg", "Sõdurioskuste kursus", "Juhtimisoskused koju", "Iga KL liige = juht", "10 000+ juhti", "Tasuta koolitus"),
    ("Partner", "KOV", "Otsused lähedal", "Kooli aiad, LEADER", "Kõik KOV kaasatud", "Puukoolid"),
    ("Partner", "ERMA / farmer", "Pakkumine", "Kohalik toode", "30 puukooli", "Hemp Authority"),
    ("Partner", "Tarbija", "Nõudlus", "Osta Eesti", "Reguleeritud turg", "Inimeste Fond"),
    ("", "", "", "", "", ""),
    ("PILOOT 2026", "Sakala malev", "Viljandi, Tactical Foodpack", "Puukoolid", "3 maleva", "Sverre võrgustik"),
    ("", "Jõgeva malev", "Sadala Agro, kanep", "Farmerid", "", "Nordic Hemp"),
    ("", "Tartu malev", "EPM, teadus", "Koolitus", "", "ERMA klaster"),
]
write_rows(ws12, 3, oless)
style_header(ws12, 3, 6, fill=PatternFill("solid", fgColor="1565C0"))
style_table(ws12, 4, 14, 6)

# Malevapealiku ülesanded
ws12.cell(row=16, column=1, value="MALEVAPEALIKU ÜLEANDED").font = BOLD
tasks = [
    ("#", "Ülesanne", "Sagedus", "Mõõdik", "Kulu", "Märkus"),
    ("1", "Kohtumine KOV juhtidega", "Kvartaalselt", "Protokoll", 0, "Järvamaa mudel"),
    ("2", "Tsiviil-maastiku kaart", "1×/a", "Farmerid, poed, koolid", 0, "Anonüümne"),
    ("3", "Nõudluse raport ERMA-le", "Kvartaalselt", "Osta Eesti statistika", 0, "Maleva piirkond"),
    ("4", "Kriisikoolitus (CIMIC)", "1×/a", "Osalejad", "LEADER/KL", "Partnerid kaasatud"),
    ("5", "Noorte aiaprogramm", "Pidev", "Kotkad/Kodutütarde", "KOV", "Loodus kooli"),
    ("6", "JUP suunamine tsiviilile", "Pidev", "Juhtide arv", 0, "Sõdurioskuste export"),
]
write_rows(ws12, 17, tasks)
style_header(ws12, 17, 6, fill=GREEN)
style_table(ws12, 18, 24, 6)

# JUP pipeline
ws12.cell(row=26, column=1, value="JUP → TSIVIIL PIPELINE").font = BOLD
jup = [
    ("Kursus", "Mida annab", "Kuhu liigub", "Näide", "Kool", "A5 maht"),
    ("Sõdurioskused / alus", "Meeskond, distsipliin", "Iga kodukohas", "KL liige", "KL kool", "10 000+"),
    ("JUP 1", "Juhtimise baas", "Meeskonnajuht", "Malevkond", "Alu mõis", "2 000+"),
    ("JUP 2", "Probleemi lahendus", "KOV kriisikomisjon", "Järvamaa", "Alu mõis", "500+"),
    ("JUP 3", "Ülesandekeskne juht", "Malevapealik", "15 maleva", "Alu mõis", "50+"),
    ("CIMIC koolitus", "Tsiviil-sõjaline", "Omavalitsus + partner", "Kriis", "Õõlg", "15 KOV/a"),
]
write_rows(ws12, 27, jup)
style_header(ws12, 27, 6)
style_table(ws12, 28, 33, 6)

# ═══════════════════════════════════════════════════════════
# LEHT 13: FINANTSSTRATEEGIA — DALIO / ROBBINS
# ═══════════════════════════════════════════════════════════
ws13 = wb.create_sheet("Finantsstrateegia")
add_title(ws13, "EESTI FINANTSSTRATEEGIA — 8 MITTE-KORRELATSIOONSET VOOGU (RAY DALIO / TONY ROBBINS)", 6)
set_widths(ws13, [8, 28, 22, 12, 14, 38])

ws13.merge_cells("A2:F2")
ws13.cell(
    row=2,
    column=1,
    value=(
        "Põhimõte — Ray Dalio (Bridgewater): leia mitte-korrelatsioonseid tulovooge ja investeeri neisse "
        "(täisversioon 15–20). Eesti mudel: 8 voogu = macro 40/30/20/10. "
        "Tony Robbins: *Money: Master the Game*, All Seasons portfell."
    ),
).alignment = Alignment(wrap_text=True)

fin_strat = [
    ("#", "Tuluvoog / varaklass", "Korrelatsioon", "Siht %", "A5 € (84M)", "Märkus"),
    ("1", "Päikesepargid + salvestus", "Energia ≠ tarbimine", 0.40, "=D4*84000000", "8× 5MW parki"),
    ("2", "Liisingumasinad + robotid", "Tootlik vara ≠ börs", 0.10, "=D5*84000000", "Paid off → vaba kassa"),
    ("3", "Tööstuskanep + eksport + CBD", "Ekspordi tsükkel", 0.10, "=D6*84000000", "Kiud, kosmeetika"),
    ("4", "Growshop + reguleeritud turg", "Kohalik tarbimine", 0.10, "=D7*84000000", "Asendab importi"),
    ("5", "Puukoolid + maaelu (LEADER)", "Maaelu ≠ börs", 0.05, "=D8*84000000", "30 puukooli"),
    ("6", "Haridus + kutsekool", "Inimkapital", 0.05, "=D9*84000000", "Uus eriala"),
    ("7", "Mirror säästud + teadus (EPM)", "Bürokraatia + pikaajaline", 0.05, "=D10*84000000", "~€35M/a + patendid"),
    ("8", "Basic income + kriisivarud", "Sotsiaal, defensiivne", 0.20, "=D11*84000000", "10k × 200€ + varud"),
    ("", "KOKKU (8 voogu)", "Macro 40/30/20/10", "=SUM(D4:D11)", "=SUM(E4:E11)", "100% Inimeste Fond A5"),
]
write_rows(ws13, 4, fin_strat)
style_header(ws13, 4, 6, fill=PatternFill("solid", fgColor="6A1B9A"))
style_table(ws13, 5, 13, 6)
for r in range(5, 12):
    ws13.cell(row=r, column=4).number_format = PCT
    ws13.cell(row=r, column=5).number_format = EUR

wb.save(OUTPUT)
print(f"Salvestatud: {OUTPUT}")
